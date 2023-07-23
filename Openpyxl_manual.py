# openpyxl 모듈 import
import openpyxl as op

# 새로운 Workbook 객체 생성
wb = op.Workbook()

# 엑셀파일 저장 - workbook 객체인 'wb'를 'openpyxl_test.xlsx' 라는 파일명으로 저장
#wb.save(./excel/openpyxl_test.xlsx")

# 엑셀파일 불러오기
path = "./excel/openpyxl_test.xlsx"
wb = op.load_workbook(path)

# wb 객체 통해 새로운 시트 생성(시트명 : new_sheet1)
ws = wb.create_sheet("new_sheet1")

# 해당 워크북(엑셀파일) 저장하기
#wb.save("./excel/test_result.xlsx")

# 활성화되어있는 시트 설정
ws = wb.active
print(ws)

# 시트명으로 접근 - sheet2로 설정하기
#ws = wb["sheet2"]

'''
# 시트명 만들기
wb = op.load_workbook("./excel/openpyxl_test.xlsx")
    # 해당 Workbook의 시트 목록을 리스트로 저장
ws_list = wb.sheetnames
print(ws_list) #리스트 출력

for sht in ws_list:
    ws = wb[sht] # Sheet 객체 생성
    print(ws) # 객체 출력
'''

# Cell의 Data 읽기 - 각 표현법의 뒤에 .value를 붙여줘야 실제값을 읽어올 수 있음에 주의!
    # 방법 1 : Sheet의 Cell 속성 사용하기
#data1 = ws.cell(row=1, column=2).value

    # 방법 2 : 엑셀 인덱스(Range) 사용하기
#data2 = ws["B1"].value

    # Cell 여러 범위로 읽어오기
#rng = ws["A1:B1"] #A1:B1 범위 저장
#print("Range(a1:b1) : ", rng) #출력값은 튜플

    # Cell 여러 범위로 읽어오기
#for  rng_data  in  rng: # 열 먼저 읽어오기
#    for  cell_data  in  rng_data: # 행 읽어오기
